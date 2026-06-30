import pickle
import pandas as pd

from openpyxl.formatting.rule import CellIsRule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from core.resources import resource_path
from core.xct_mapping import COL_MAP_XCT1, COL_MAP_XCT2

PRED_COLS = ['tt.bmd', 'tt.ar', 'tb.bmd', 'bv/tv', 'tb.n', 'tb.th', 'tb.sp', 'tb.1/n.sd',
             'tb.ar', 'ct.bmd', 'ct.th', 'ct.po', 'ct.po.dm', 'ct.pm', 'ct.ar']

def load_scaler(xct_gen):
    scalers = {}
    radius_path = resource_path(f"models/radius_XCT{xct_gen}_scaler.pkl")
    with open(radius_path, 'rb') as f:
        scalers['radius'] = pickle.load(f)

    tibia_path = resource_path(f"models/tibia_XCT{xct_gen}_scaler.pkl")
    with open(tibia_path, 'rb') as f:
        scalers['tibia'] = pickle.load(f)
    return scalers

def load_model(xct_gen, model_type):
    machine = "old" if xct_gen == 1 else "new"
    weighted = "_balanced" if model_type == "balanced" else ""

    models = {}

    radius_path = resource_path(f"models/radius_{machine}{weighted}_model.pkl")
    with open(radius_path, 'rb') as f:
        models['radius'] = pickle.load(f)

    tibia_path = resource_path(f"models/tibia_{machine}{weighted}_model.pkl")
    with open(tibia_path, 'rb') as f:
        models['tibia'] = pickle.load(f)

    return models

def conformal_marking(output_path, dataframe, highlight):

    # get the cell range
    confidence_col_index = dataframe.columns.get_loc("Confidence") + 1
    confidence_col_letter = get_column_letter(confidence_col_index)
    max_row = len(dataframe)+1
    cell_range = f"{confidence_col_letter}2:{confidence_col_letter}{max_row}"

    # process the workbook
    wb = load_workbook(output_path)
    ws = wb.active

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    if "85th" in highlight:
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator="lessThan", formula=["0.61"], fill=red_fill))
    if "95th" in highlight:
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator="between", formula=["0.61", "0.8"], fill=yellow_fill))

    wb.save(output_path)

def save_excel(dataframe, file_path, sheet_name):
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

def run_processing(dataframe: pd.DataFrame, model_type, xct_gen, conformal, output_path, sheet_name):

    # remove XCT2 specific columns from pred_cols if not needed
    selected_cols = PRED_COLS.copy()
    col_map = COL_MAP_XCT2
    if xct_gen == 1:
        selected_cols.remove('ct.po')
        selected_cols.remove('ct.po.dm')
        col_map = COL_MAP_XCT1
    
    original_cols = dataframe.columns.copy()

    dataframe.rename(columns=col_map, inplace=True)
    dataframe.columns = dataframe.columns.str.lower()

    # load needed models
    scalers = load_scaler(xct_gen)
    models = load_model(xct_gen, model_type)

    output_preds = []
    output_probs = []

    for _, row in dataframe.iterrows():
        if pd.isna(row["site"]):
            output_preds.append("invalid")
            output_probs.append(-1)
            continue # skip rows without measurement data
        if row['site'].startswith('R'):
            site = 'radius'
        elif row['site'].startswith('T'):
            site = 'tibia'
        else:
            raise Exception(f"Unknown site {row['site']}: Make sure Site column indicates Radius (R) or Tibia (T).")

        values = [row[selected_cols].fillna(0).tolist()]
        values_transformed = scalers[site].transform(values)
        prediction = models[site].predict(values_transformed)
        prediction_prob = models[site].predict_proba(values_transformed)

        output_preds.append("pass" if prediction[0] == 0 else "fail")
        output_probs.append(round(max(prediction_prob[0]), 3))

    dataframe['Grading'] = output_preds
    dataframe['Confidence'] = output_probs
    dataframe.columns = list(original_cols) + ["Grading", "Confidence"]

    save_excel(dataframe, output_path, sheet_name)

    if len(conformal) > 0:
        conformal_marking(output_path, dataframe, conformal)

    return
